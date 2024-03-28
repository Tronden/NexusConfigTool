import tkinter as tk
from tkinter import messagebox, ttk, PhotoImage
import os
import sys
import shutil
import ctypes as ct
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Fill, Alignment

data_folder = 'Data'

class ExcelCreationToolGUI:
    def __init__(self, root):
        # Determine if running as a script or frozen executable
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Running as a bundled executable
            self.base_dir = sys._MEIPASS
        else:
            # Running as a normal script
            self.base_dir = os.path.dirname(__file__)
        
        self.image_path = os.path.join(self.base_dir, "Data", "GUI")
        icon_path = os.path.join(self.base_dir, "Data", "GUI", "FM_icon.ico")
        
        self.font = "Arial"
        self.font_size = 12
        self.is_dark_mode = True  
        self.active_ems_plc = None
        self.active_ess_plc = None


        self.root = root
        
        root.title("Nexus Config Tool")
        root.iconbitmap(icon_path)
        self.dark_title_bar(root)
        self.setup_ui(self.root)

    def setup_ui(self, root):
        self.light = PhotoImage(file=self.image_path+"/light.png")
        self.dark = PhotoImage(file=self.image_path+"/dark.png")
        
        self.light_mode = {
            "bg": "white", 
            "fg": "black",
            "frame": {"bg": "white", "fg": "black"},
            "label": {"bg": "white", "fg": "black"},
            "entry": {"bg": "white", "fg": "black"},
            "button": {"bg": "white", "fg": "black"},
            "combobox": {"bg": "white", "fg": "black"},
        }

        self.dark_mode = {
            "bg": "#3c3e3c",
            "fg": "white",
            "frame": {"bg": "#3c3e3c", "fg": "white"},
            "label": {"bg": "#3c3e3c", "fg": "white"},
            "entry": {"bg": "#3c3e3c", "fg": "black"},
            "button": {"bg": "#3c3e3c", "fg": "black"},
            "combobox": {"bg": "#3c3e3c", "fg": "black"},
        }

        self.style = ttk.Style()
    
        self.apply_theme(self.dark_mode)

        main_frame = ttk.Frame(root)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=20)
        
        # Toggle Dark Mode
        self.theme_button = ttk.Label(main_frame, image=self.dark, padding=(0,0,0,20), cursor="hand2")
        self.theme_button.grid(row=0, column=0, sticky=tk.W)
        self.theme_button.bind("<Button-1>", self.toggle_theme)
        
        # Header Image
        self.logo = PhotoImage(file=self.image_path+"/FM.png")
        logo_label = ttk.Label(main_frame, image=self.logo, padding=(0,0,0,20)).grid(row=1, column=0, columnspan=2)
        
        # Barge Number
        self.barge_number_label = ttk.Label(main_frame, text="Barge Number :").grid(row=2, column=0, sticky=tk.W)
        self.barge_number_entry = ttk.Entry(main_frame)
        self.barge_number_entry.grid(row=2, column=1, sticky=tk.EW)
       

        # Fjord Control Password
        ttk.Label(main_frame, text="FC Password :").grid(row=3, column=0, sticky=tk.W)
        self.fjord_control_password_entry = ttk.Entry(main_frame)
        self.fjord_control_password_entry.grid(row=3, column=1, sticky=tk.EW)

        # Send Interval
        ttk.Label(main_frame, text="Send Interval:").grid(row=4, column=0, sticky=tk.W)
        self.send_interval_entry = ttk.Entry(main_frame)
        self.send_interval_entry.grid(row=4, column=1, sticky=tk.EW)
        self.send_interval_entry.insert(0, "5000")

        # EMS PLC Type Buttons Frame
        self.ems_plc_type_var = tk.StringVar()
        self.ess_plc_type_var = tk.StringVar()
        self.create_plc_type_buttons(main_frame, "EMS Type :", ["Beckhoff", "Wago"], self.ems_plc_type_var, 5)
        self.create_plc_type_buttons(main_frame, "ESS Type :", ["Beckhoff", "Wago"], self.ess_plc_type_var, 6)

        # Number of Generators
        ttk.Label(main_frame, text="Number of Generators :").grid(row=7, column=0, sticky=tk.W)
        self.num_generators_combobox = ttk.Combobox(main_frame, values=[str(i) for i in range(1, 4)], state='readonly')
        self.num_generators_combobox.grid(row=7, column=1, sticky=tk.EW)
        self.num_generators_combobox.bind("<<ComboboxSelected>>", self.show_gen_settings)

        # Generator Settings Placeholder
        self.gen_settings_frame = ttk.Frame(main_frame, padding="20")
        self.gen_settings_frame.grid(row=8, column=0, columnspan=2, sticky=tk.EW)

        # Create File Button
        self.create_file_button = ttk.Button(main_frame, text="Create File", command=self.create_file).grid(row=9, column=0, columnspan=2, sticky=tk.EW)
   
    def set_plc_type(self, plc_category, plc_type):
        if plc_category == "ems":
            self.ems_plc_type = plc_type
            self.update_button_state(self.ems_plc_buttons, plc_type)
        elif plc_category == "ess":
            self.ess_plc_type = plc_type
            self.update_button_state(self.ess_plc_buttons, plc_type)

    def create_plc_type_buttons(self, parent, label, options, variable, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky=tk.W)
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=1, sticky=tk.EW)
        for option in options:
            rb = ttk.Radiobutton(frame, text=option, value=option, variable=variable)
            rb.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    def dark_title_bar(self, window):
        window.update()
        DWMWA_USE_IMMERSIVE_DARK_MODE = 20
        set_window_attribute = ct.windll.dwmapi.DwmSetWindowAttribute
        get_parent = ct.windll.user32.GetParent
        hwnd = get_parent(window.winfo_id())
        rendering_policy = DWMWA_USE_IMMERSIVE_DARK_MODE
        value = 2
        value = ct.c_int(value)
        set_window_attribute(hwnd, rendering_policy, ct.byref(value), ct.sizeof(value))

    def get_theme(self):
        return self.dark_mode if self.is_dark_mode else self.light_mode

    def toggle_theme(self, event=None):
        # Toggle the dark mode flag first
        self.is_dark_mode = not self.is_dark_mode
        # Apply the theme based on the new state
        theme = self.dark_mode if self.is_dark_mode else self.light_mode
        self.apply_theme(theme)
        # Update the theme button image based on the current mode
        self.theme_button.config(image=self.dark if self.is_dark_mode else self.light)
       
    def apply_theme(self, theme):
        self.root.config(bg=theme["bg"])
        # Configure global style for all widgets
        self.style.configure('TLabel', background=theme["label"]["bg"], foreground=theme["label"]["fg"], font=(self.font, self.font_size), padding=5)
        self.style.configure('TEntry', background=theme["entry"]["bg"], foreground=theme["entry"]["fg"], font=(self.font, self.font_size), insertbackground=theme["entry"]["fg"], padding=5)
        self.style.configure('TButton', font=('Arial', 12), padding=5)
        self.style.configure('TFrame', background=theme["frame"]["bg"], foreground=theme["frame"]["fg"])
        self.style.configure('TLabelframe', background=theme["frame"]["bg"], foreground=theme["frame"]["fg"], font=(self.font, self.font_size))
        self.style.configure('TLabelframe.Label', background=theme["frame"]["bg"], foreground=theme["frame"]["fg"], font=(self.font, self.font_size))
        self.style.configure('TCombobox', background=theme["combobox"]["bg"], foreground=theme["combobox"]["fg"], font=(self.font, self.font_size), padding=5)
        self.style.configure('TRadiobutton', background=theme["frame"]["bg"], foreground=theme["frame"]["fg"], font=(self.font, self.font_size), padding=5)
        # Update all widgets with new theme (this is an example and needs to be adapted to your actual widgets)
        for widget in self.root.winfo_children():
            self.update_widget(widget)

    def update_widget(self, widget):
        # Apply theme recursively for container widgets
        if isinstance(widget, ttk.Frame):
            widget.config(style='TFrame')
        elif isinstance(widget, ttk.Labelframe):
            widget.config(style='TLabelframe')
        elif isinstance(widget, ttk.Label):
            widget.config(style='TLabel')
        elif isinstance(widget, ttk.Entry):
            widget.config(style='TEntry')
        elif isinstance(widget, ttk.Button):
            widget.config(style='TButton')
            
    def show_gen_settings(self, event=None):
        for widget in self.gen_settings_frame.winfo_children():
            widget.destroy()

        self.gen_settings = []
        num_generators = int(self.num_generators_combobox.get())
        
        for i in range(num_generators):
            settings_frame = ttk.LabelFrame(self.gen_settings_frame, text=f"Generator {i + 1}")
            settings_frame.grid(row=0, column=i*2, sticky="nw", padx=5, pady=5, columnspan=2)
            
            # Panel Type
            ttk.Label(settings_frame, text="Panel Type :").grid(row=0, column=0, sticky="w")
            panel_type_combobox = ttk.Combobox(settings_frame, values=["No Panel", "DSE 8610 MKII", "InteliLite 4 AMF 25", "Sices GC600"], state='readonly')
            panel_type_combobox.grid(row=0, column=1, columnspan=2, padx=5, pady=5)
            
            # Communication Type
            ttk.Label(settings_frame, text="Communication Type :").grid(row=1, column=0, sticky="w")
            com_type_combobox = ttk.Combobox(settings_frame, values=["RTU", "TCP"], state='readonly')
            com_type_combobox.grid(row=1, column=1, columnspan=2, padx=5, pady=5)

            # Dynamic settings frame
            dynamic_settings_frame = ttk.Frame(settings_frame)
            dynamic_settings_frame.grid(row=2, column=0, columnspan=2, sticky="ew")

            self.gen_settings.append({
                "panel_type": panel_type_combobox,
                "com_type": com_type_combobox,
                "dynamic_settings_frame": dynamic_settings_frame,
            })
            com_type_combobox.bind("<<ComboboxSelected>>", lambda event, index=i: self.on_com_type_selected(event, index))
            self.update_widget(settings_frame)

    def on_com_type_selected(self, event, index):
        com_type = self.gen_settings[index]["com_type"].get()
        dynamic_settings_frame = self.gen_settings[index]["dynamic_settings_frame"]
        # Clear previous content
        for widget in dynamic_settings_frame.winfo_children():
            widget.destroy()
        if com_type == "RTU":
            self.show_rtu_settings(dynamic_settings_frame, index)
        elif com_type == "TCP":
            self.show_tcp_settings(dynamic_settings_frame, index)

    def show_rtu_settings(self, settings_frame, index):
        # Clear existing widgets in this frame
        for widget in settings_frame.winfo_children():
            widget.destroy()
        
        # Slave Address Combobox
        ttk.Label(settings_frame, text="Slave Address :").grid(row=1, column=0, sticky='w')
        self.slave_address_combobox = ttk.Combobox(settings_frame, values=[str(n) for n in range(1, 31)], state='readonly')
        self.slave_address_combobox.set("10")  # Default value
        self.slave_address_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.slave_address_combobox.bind("<<ComboboxSelected>>", lambda event, l="Slave Address", cb=self.slave_address_combobox, i=index: self.on_combobox_change(l, cb.get(), i))
    
        # Com Port Combobox
        ttk.Label(settings_frame, text="Com Port :").grid(row=2, column=0, sticky='w')
        self.com_port_combobox = ttk.Combobox(settings_frame, values=[f"COM{n}" for n in range(1, 5)], state='readonly')
        self.com_port_combobox.set("COM1")  # Default value
        self.com_port_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.com_port_combobox.bind("<<ComboboxSelected>>", lambda event, l="Com Port", cb=self.com_port_combobox, i=index: self.on_combobox_change(l, cb.get(), i))

        # Baudrate Combobox
        ttk.Label(settings_frame, text="Baudrate :").grid(row=3, column=0, sticky='w')
        self.baudrate_combobox = ttk.Combobox(settings_frame, values=["9600", "19200", "115200"], state='readonly')
        self.baudrate_combobox.set("19200")  # Default value
        self.baudrate_combobox.grid(row=3, column=1, padx=5, pady=5)
        self.baudrate_combobox.bind("<<ComboboxSelected>>", lambda event, l="Baudrate", cb=self.baudrate_combobox, i=index: self.on_combobox_change(l, cb.get(), i))
        
        # Stopbit Combobox
        ttk.Label(settings_frame, text="Stopbit :").grid(row=4, column=0, sticky='w')
        self.stopbit_combobox = ttk.Combobox(settings_frame, values=["1", "1.5", "2"], state='readonly')
        self.stopbit_combobox.set("2")  # Default value
        self.stopbit_combobox.grid(row=4, column=1, padx=5, pady=5)
        self.stopbit_combobox.bind("<<ComboboxSelected>>", lambda event, l="Stopbit", cb=self.stopbit_combobox, i=index: self.on_combobox_change(l, cb.get(), i))
        
        # Parity Combobox
        ttk.Label(settings_frame, text="Parity :").grid(row=5, column=0, sticky='w')
        self.parity_combobox = ttk.Combobox(settings_frame, values=["None", "Odd", "Even"], state='readonly')
        self.parity_combobox.set("None")  # Default value
        self.parity_combobox.grid(row=5, column=1, padx=5, pady=5)
        self.parity_combobox.bind("<<ComboboxSelected>>", lambda event, l="Parity", cb=self.parity_combobox, i=index: self.on_combobox_change(l, cb.get(), i))
        
        # Databit Combobox
        ttk.Label(settings_frame, text="Databit :").grid(row=6, column=0, sticky='w')
        self.databit_combobox = ttk.Combobox(settings_frame, values=["7", "8"], state='readonly')
        self.databit_combobox.set("8")  # Default value
        self.databit_combobox.grid(row=6, column=1, padx=5, pady=5)
        self.databit_combobox.bind("<<ComboboxSelected>>", lambda event, l="Databit", cb=self.databit_combobox, i=index: self.on_combobox_change(l, cb.get(), i))

        self.gen_settings[index]["settings"] = {
            "Slave Address": self.slave_address_combobox.get(),
            "Com Port": self.com_port_combobox.get(),
            "Baudrate": self.baudrate_combobox.get(),
            "Stopbit": self.stopbit_combobox.get(),
            "Parity": self.parity_combobox.get(),
            "Databit": self.databit_combobox.get()
        }
        
    def show_tcp_settings(self, settings_frame, index):
        # Clear existing widgets in this frame
        for widget in settings_frame.winfo_children():
            widget.destroy()

        # Slave Address Combobox
        ttk.Label(settings_frame, text="Slave Address :").grid(row=1, column=0, sticky='w')
        self.slave_address_combobox = ttk.Combobox(settings_frame, values=[str(n) for n in range(1, 31)], state='readonly')
        self.slave_address_combobox.set("10")  # Default value
        self.slave_address_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.slave_address_combobox.bind("<<ComboboxSelected>>", lambda event, l="Slave Address", cb=self.slave_address_combobox, i=index: self.on_combobox_change(l, cb.get(), i))
    
        
        ttk.Label(settings_frame, text="IP Address :").grid(row=2, column=0, sticky='w')
        self.ip_address_entry = ttk.Entry(settings_frame)
        self.ip_address_entry.insert(0,"192.168.1.110")
        self.ip_address_entry.grid(row=2, column=1, padx=5, pady=5)
        self.ip_address_entry.bind("<KeyRelease>", lambda event, l="IP Address", entry=self.ip_address_entry, i=index: self.on_entry_change(l, entry.get(), i))

        ttk.Label(settings_frame, text="Port :").grid(row=3, column=0, sticky='w')
        self.port_entry = ttk.Entry(settings_frame)
        self.port_entry.insert(0,"502")
        self.port_entry.grid(row=3, column=1, padx=5, pady=5)
        self.port_entry.bind("<KeyRelease>", lambda event, l="Port", entry=self.port_entry, i=index: self.on_entry_change(l, entry.get(), i))


        self.gen_settings[index]["settings"] = {
            "Slave Address": self.slave_address_combobox.get(),
            "IP Address": self.ip_address_entry.get(),
            "Port": self.port_entry.get(),
        }
    
    def on_combobox_change(self, label_text, value, index):
        self.gen_settings[index]["settings"][label_text] = value  
         
    def on_entry_change(self, label_text, value, index):
        self.gen_settings[index]["settings"][label_text] = value
         
    def create_file(self):
        barge_number = self.barge_number_entry.get()
        fc2_barge_number = barge_number.zfill(3)
        fjord_control_password = self.fjord_control_password_entry.get()
        send_interval = self.send_interval_entry.get()
        ems_plc_type = self.ems_plc_type_var.get()
        ess_plc_type = self.ess_plc_type_var.get()
        num_generators =  int(self.num_generators_combobox.get())
        panelmodifications = {}
        panels = {}
        rtumodifications = {}
        tcpmodifications = {}
        next_tcp = 1
        
        #Directory setup
        base_dir = f"Configs/FH{barge_number}"
        modules_dir = os.path.join(base_dir, "Modules")
        os.makedirs(modules_dir, exist_ok=True)
        
        #Copying the correct modules
        shutil.copy(f"Data/Modules/FCTags 3.0 {ems_plc_type}.nxbin", modules_dir)
        shutil.copy(f"Data/Modules/Effektberegninger.nxbin", modules_dir)
        shutil.copytree("Data/Modules/Data", os.path.join(modules_dir, "Data"))
       
        # Creating the Excel workbook
        config_excel_path = os.path.join(base_dir, "config.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        self.copy_sheet_from_template(
            os.path.join("Data", "FC2", "Fjord Control 2.xlsx"),wb,"Fjord Control 2",modifications={"C8": f"FH{fc2_barge_number}"}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "EMS", f"EMS modbus {ems_plc_type}.xlsx"),wb,"EMS modbus",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "EXT", "Ark1.xlsx"),wb,"Ark1",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "ESS", f"ESS modbus {ess_plc_type}.xlsx"),wb,"ESS modbus",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "Alarmlist", f"Alarmlist {num_generators}.xlsx"),wb,"AlarmList",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "Fjord Control Online", f"Fjord Control Online {num_generators}.xlsx"),wb,"Fjord Control Online",modifications={"C5":f"bearer FH{barge_number}:{fjord_control_password}","C6":f"FH{barge_number}","C7":float(send_interval)}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "Fjord Control Online Cloud", f"Fjord Control Online Cloud {num_generators}.xlsx"),wb,"Fjord Control Online Cloud",modifications={"C5":f"bearer FH{barge_number}:{fjord_control_password}","C6":f"FH{barge_number}","C7":float(send_interval)}, namechange = None
        )
        
        for i in range(num_generators):
            panel_type = self.gen_settings[i]["panel_type"].get()
            com_type = self.gen_settings[i]["com_type"].get()
        
            if panel_type != "No Panel":
                settings = self.gen_settings[i]["settings"]
                com_port_column_map = {'COM1':'C','COM2':'F','COM3':'I','COM4':'L',}
                generator_ip_column_map = {'1':'I','2':'L','3':'O',}
                generator_ip_letter_column_map = {'1':'G','2':'J','3':'M',}

                if com_type == "RTU":
                    slave_address = settings.get("Slave Address")
                    panelmodifications["C3"] = f"Generator{i+1}" if com_type == "TCP" else f"{settings.get('Com Port')}"
                    panelmodifications["D6"] = f"{com_type}"
                    panelmodifications["D7"] = 1000
                    panelmodifications[f"A1{i+1}"] = slave_address
                    panelmodifications[f"B1{i+1}"] = "Modbus"
                    panelmodifications[f"C1{i+1}"] = "Modbus"
                    panelmodifications[f"D1{i+1}"] = "Modbus"

                    if panel_type in ["DSE 8610 MKII", "InteliLite 4 AMF 25"]:
                        for idx in range(row, row+28):
                            panelmodifications[f"F{idx}"] = slave_address
                        row = row+28

                    elif panel_type == "Sices GC600":
                        for idx in range(row, row+166):
                            panelmodifications[f"F{idx}"] = slave_address
                        row = row+166

                    panels[i][row] = row

                    com_port = comsettings.get("Com Port")
                    if com_port in com_port_column_map:
                        base_column = com_port_column_map[com_port]
                        # Directly update the main rtumodifications dictionary
                        rtumodifications[f"{base_column}{5}"] = com_port
                        rtumodifications[f"{base_column}{6}"] = int(comsettings.get("Baudrate"))
                        rtumodifications[f"{base_column}{7}"] = float(comsettings.get("Stopbit"))
                        rtumodifications[f"{base_column}{8}"] = comsettings.get("Parity")
                        rtumodifications[f"{base_column}{9}"] = int(comsettings.get("Databit"))

                elif com_type == "TCP":
                    panelmodifications[sheet_key]["C1"] = f"Modbus {com_type} Master"
                    panelmodifications[sheet_key]["C3"] = f"Generator{i+1}" if com_type == "TCP" else f"{settings.get('Com Port')}"
                    panelmodifications[sheet_key]["D6"] = f"{com_type}"
                    panelmodifications[sheet_key]["D7"] = 1000
                    panelmodifications[sheet_key]["A11"] = slave_address
                    panelmodifications[sheet_key]["B11"] = "Modbus"
                    panelmodifications[sheet_key]["C11"] = "Modbus"
                    panelmodifications[sheet_key]["D11"] = "Modbus"
                    if panel_type in ["DSE 8610 MKII", "InteliLite 4 AMF 25"]:
                        for idx in range(row, row+28):
                            modifications[f"F{idx}"] = slave_address
                    elif panel_type == "Sices GC600":
                        for idx in range(row, row+166):
                            modifications[f"F{idx}"] = slave_address

                    Generator = f"Generator{i+1}"
                    base_column = generator_ip_column_map[str(next_tcp)]
                    base_column_letter = generator_ip_letter_column_map[str(next_tcp)]
                    # Directly update the main tcpmodifications dictionary
                    tcpmodifications[f"{base_column_letter}{5}"] = "Connector name"
                    tcpmodifications[f"{base_column_letter}{6}"] = "Port no."
                    tcpmodifications[f"{base_column_letter}{7}"] = "IP"
                    tcpmodifications[f"{base_column}{5}"] = Generator
                    tcpmodifications[f"{base_column}{6}"] = int(comsettings.get("Port"))
                    tcpmodifications[f"{base_column}{7}"] = comsettings.get("IP Address")
                    next_tcp += 1
        
        if rtumodifications:
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "COM Ports.xlsx"), wb, "COM Ports", modifications=rtumodifications, namechange=None
            )
        else:
               # If no RTU settings found, copy the sheet without modifications
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "COM Ports.xlsx"), wb, "COM Ports", modifications={}, namechange=None
            )

        if tcpmodifications:
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "TCP Ports.xlsx"), wb, "TCP Ports", modifications=tcpmodifications, namechange=None
            )
        else:
               # If no TCP settings found, copy the sheet without modifications
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "TCP Ports.xlsx"), wb, "TCP Ports", modifications={}, namechange=None
            )
    
        self.copy_sheet_from_template(
           os.path.join("Data", "EXT", "Error values.xlsx"),wb,"Error values",modifications={}, namechange = None
        )
            
        for i in sheet_key:
            self.copy_sheet_from_template(
                os.path.join("Data", "Panels", f"{panel_type}.xlsx"), wb, "Generator", modifications=modifications, namechange=sheet_name
            )
            for i in panel_type[sheet_key]:
                row = sheet_row[sheet_key][i]
                self.copy_range_to_sheet(os.path.join("Data", "Panels", f"{panel_type}.xlsx"), "Generator", "A1:G28", wb, sheet_key, f"F{row}")
        
        # Save the new config
        self.save(config_excel_path, wb)


    def aggregate_generator_settings(self):
        self.aggregated_modifications = {}
        for i, gen_settings in enumerate(self.gen_settings):
            com_type = gen_settings["com_type"].get()
            panel_type = gen_settings["panel_type"].get()
            settings = gen_settings["settings"]
            sheet_key = self.get_sheet_key(com_type, settings, i)

            if sheet_key not in self.aggregated_modifications:
                self.aggregated_modifications[sheet_key] = {
                    "com_type": com_type,
                    "generators": []
                }
            
            self.aggregated_modifications[sheet_key]["generators"].append({
                "panel_type": panel_type,
                "settings": settings,
                "index": i
            })

    def get_sheet_key(self, com_type, settings, index):
        if com_type == "RTU":
            return f"Generator {com_type} {settings.get('Com Port')}"
        else:
            return f"Generator {com_type} {index + 1}"
            
        
    def copy_sheet_from_template(self, template_path, new_wb, original_sheet_name, modifications=None, namechange=None):
        # Load the template workbook
        template_wb = openpyxl.load_workbook(template_path, data_only=True)

        # Check if the original sheet exists in the template
        if original_sheet_name not in template_wb.sheetnames:
            print(f"Sheet '{original_sheet_name}' not found in {template_path}.")
            return

        # Get the template sheet
        template_sheet = template_wb[original_sheet_name]

        # Determine the new sheet name: use 'namechange' if provided, otherwise use the original sheet name
        new_sheet_name = namechange if namechange else original_sheet_name

        # Check if the new sheet name already exists in the target workbook to prevent overwriting
        if new_sheet_name in new_wb.sheetnames:
            print(f"Sheet name '{new_sheet_name}' already exists in the workbook. Please use a unique name.")
            return

        # Create a new sheet in the target workbook with the new sheet name
        new_sheet = new_wb.create_sheet(title=new_sheet_name)

        # Copy cells from template sheet to the new sheet, including values and styles
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Apply any modifications specified
        if modifications:
            for cell, value in modifications.items():
                new_sheet[cell].value = value

        # Copy merged cell ranges from the template sheet to the new sheet
        for merged_cell_range in template_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_cell_range))
        
        self.auto_adjust_column_widths(new_sheet)
        
        template_wb.close()
        
    def auto_adjust_column_widths(self, sheet):
        column_widths = {}
        default_width = 8.43
        padding = 0.2
        average_char_width = 1.2

        # Collect all merged cells in a set for quick lookup
        merged_cells = set()
        for merged_cell_range in sheet.merged_cells.ranges:
            for row in sheet.iter_rows(min_row=merged_cell_range.min_row, max_row=merged_cell_range.max_row,
                                    min_col=merged_cell_range.min_col, max_col=merged_cell_range.max_col):
                for cell in row:
                    merged_cells.add(cell.coordinate)

        # Handle normal cells: calculate max width for each column
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and cell.coordinate not in merged_cells:  # Skip if cell is merged
                    column_letter = get_column_letter(cell.column)
                    calculated_width = (len(str(cell.value)) * average_char_width) + padding
                    column_widths[column_letter] = max(column_widths.get(column_letter, default_width), calculated_width)

        # Adjust for merged cells, focusing only on adjusting the first column of the merged range
        for merged_cell_range in sheet.merged_cells.ranges:
            first_cell = sheet.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col)
            span_length = merged_cell_range.max_col - merged_cell_range.min_col
            if first_cell.value:
                column_letter = get_column_letter(first_cell.column)
                # Here we consider the content length of the merged cells
                required_width = len(str(first_cell.value)) + padding - (span_length * default_width)
                # Update the column width if the calculated width is greater
                column_widths[column_letter] = max(column_widths.get(column_letter, default_width), required_width)

        # Apply the calculated widths to the columns
        for column_letter, width in column_widths.items():
            final_width = max(width, default_width)
            sheet.column_dimensions[column_letter].width = final_width

    def copy_range_to_sheet(src_wb_path, src_sheet_name, src_range, tgt_wb, tgt_sheet_name, tgt_start_cell):
        """
        Copies a range from a source workbook and sheet to a specified location in a target workbook and sheet.
        
        Parameters:
        - src_wb_path: Path to the source workbook.
        - src_sheet_name: Name of the source sheet.
        - src_range: The range in A1 notation (e.g., "A1:C3") to copy from the source sheet.
        - tgt_wb: The target workbook object.
        - tgt_sheet_name: The target sheet name where the range will be pasted.
        - tgt_start_cell: The starting cell (e.g., "B4") in the target sheet where the copying range will begin.
        """
        src_wb = openpyxl.load_workbook(src_wb_path, data_only=True)
        src_sheet = src_wb[src_sheet_name]
        
        tgt_sheet = tgt_wb[tgt_sheet_name]
        
        min_col = openpyxl.utils.column_index_from_string(src_range.split(':')[0].translate({ord(c): None for c in '1234567890'}))
        min_row = int(''.join(filter(str.isdigit, src_range.split(':')[0])))
        max_col = openpyxl.utils.column_index_from_string(src_range.split(':')[1].translate({ord(c): None for c in '1234567890'}))
        max_row = int(''.join(filter(str.isdigit, src_range.split(':')[1])))
        
        tgt_start_col = openpyxl.utils.column_index_from_string(tgt_start_cell.translate({ord(c): None for c in '1234567890'}))
        tgt_start_row = int(''.join(filter(str.isdigit, tgt_start_cell)))
        
        for row in src_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                tgt_cell_row = tgt_start_row + cell.row - min_row
                tgt_cell_col = get_column_letter(tgt_start_col + cell.column - min_col)
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].value = cell.value

                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].font = copy(cell.font)
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].border = copy(cell.border)
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].fill = copy(cell.fill)
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].number_format = cell.number_format
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].protection = copy(cell.protection)
                tgt_sheet[f"{tgt_cell_col}{tgt_cell_row}"].alignment = copy(cell.alignment)

        src_wb.close()

    def save(self, config_excel_path, wb):
        try:
            wb.save(config_excel_path)
            messagebox.showinfo("Success", "Config created successfully!", parent=self.root)
            self.update_widget(self.root)
            self.root.destroy()  # Close the application after showing the success message
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save the config file.\n{e}", parent=self.root)
            self.update_widget(self.root)
                     
if __name__ == "__main__":
    # Ensure the current working directory is set to the script's or executable's directory
    root = tk.Tk()
    app = ExcelCreationToolGUI(root)
    root.mainloop()