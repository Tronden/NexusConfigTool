import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import PhotoImage
import os
import sys
import shutil
import openpyxl
from copy import copy
from openpyxl.styles import Font, Border, Fill, Alignment

data_folder = 'Data'

class ExcelCreationToolGUI:
    def __init__(self, root):
        # Determine if running as a script or frozen executable
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Running as a bundled executable
            self.base_dir = sys._MEIPASS
        else:
            # Running as a normal script
            self.base_dir = os.path.dirname(__file__)
            
        self.active_ems_plc = None
        self.active_ess_plc = None
        
        self.setup_ui(root)
        self.root = root
        
        self.gen_settings= []
        
    def setup_ui(self, root):
        root.title("Config Creation Tool")
        self.style = ttk.Style()
        self.style.configure('TFrame', padding=10)
        self.style.configure('TLabel', padding=5)
        self.style.configure('TEntry', padding=5)
        self.style.configure('TButton', padding=5)

        main_frame = ttk.Frame(root)
        main_frame.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Header Image
        image_path = os.path.join(self.base_dir, "Data", "GUI", "FM.png")
        self.image = PhotoImage(file=image_path)
        image_label = ttk.Label(main_frame, image=self.image)
        image_label.grid(row=0, column=0, columnspan=2)
        # Barge Number
        ttk.Label(main_frame, text="Barge Number:").grid(row=1, column=0, sticky=tk.W)
        self.barge_number_entry = ttk.Entry(main_frame)
        self.barge_number_entry.grid(row=1, column=1, sticky=tk.EW)

        # Fjord Control Password
        ttk.Label(main_frame, text="Fjord Control Password:").grid(row=2, column=0, sticky=tk.W)
        self.fjord_control_password_entry = ttk.Entry(main_frame)
        self.fjord_control_password_entry.grid(row=2, column=1, sticky=tk.EW)

        # Send Interval
        ttk.Label(main_frame, text="Send Interval:").grid(row=3, column=0, sticky=tk.W)
        self.send_interval_entry = ttk.Entry(main_frame)
        self.send_interval_entry.grid(row=3, column=1, sticky=tk.EW)
        self.send_interval_entry.insert(0, "2000")

        # EMS PLC Type Buttons Frame
        self.ems_plc_type_var = tk.StringVar()
        self.ess_plc_type_var = tk.StringVar()
        self.create_plc_type_buttons(main_frame, "EMS PLC Type:", ["Beckhoff", "Wago"], self.ems_plc_type_var, 4)
        self.create_plc_type_buttons(main_frame, "ESS PLC Type:", ["Beckhoff", "Wago"], self.ess_plc_type_var, 5)

        # Number of Generators
        ttk.Label(main_frame, text="Number of Generators:").grid(row=6, column=0, sticky=tk.W)
        self.num_generators_combobox = ttk.Combobox(main_frame, values=[str(i) for i in range(1, 5)], state='readonly')
        self.num_generators_combobox.grid(row=6, column=1, sticky=tk.EW)
        self.num_generators_combobox.bind("<<ComboboxSelected>>", self.show_gen_settings)

        # Generator Settings Placeholder
        self.gen_settings_frame = ttk.Frame(main_frame, padding="20")
        self.gen_settings_frame.grid(row=7, column=0, columnspan=2, sticky=tk.EW)

        # Create File Button
        self.create_file_button = ttk.Button(main_frame, text="Create File", command=self.create_file)
        self.create_file_button.grid(row=8, column=0, columnspan=2, sticky=tk.EW)
        
        ttk.Label(main_frame, text="EMS PLC Type:").grid(row=4, column=0, sticky=tk.W)
    
    def set_plc_type(self, plc_category, plc_type):
        if plc_category == "ems":
            self.ems_plc_type = plc_type
            self.update_button_state(self.ems_plc_buttons, plc_type)
        elif plc_category == "ess":
            self.ess_plc_type = plc_type
            self.update_button_state(self.ess_plc_buttons, plc_type)

    def create_plc_type_buttons(self, parent, label, options, variable, row):
        ttk.Label(parent, text=label).grid(row=row, column=0, sticky=tk.W)
        frame = ttk.Frame(parent)
        frame.grid(row=row, column=1, sticky=tk.EW)
        for option in options:
            rb = ttk.Radiobutton(frame, text=option, value=option, variable=variable)
            rb.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
    def show_gen_settings(self, event=None):
        for widget in self.gen_settings_frame.winfo_children():
                widget.destroy()
                
        num_generators = int(self.num_generators_combobox.get())
        
        for i in range(1,num_generators+1):
            settings_frame = ttk.LabelFrame(self.gen_settings_frame, text=f"Generator {i}")
            settings_frame.grid(row=0, column=i*2, sticky="ew", padx=5, pady=5, columnspan=2)

            # Panel Type
            ttk.Label(settings_frame, text="Panel Type:").grid(row=0, column=0, sticky="w")
            panel_type_combobox = ttk.Combobox(settings_frame, values=["DSE 8610 MKII", "InteliLite 4 AMF 25", "Sices GC600"], state='readonly')
            panel_type_combobox.grid(row=0, column=1, padx=5, pady=5)
    
            # Communication Type with RTU as default
            ttk.Label(settings_frame, text="Communication Type:").grid(row=1, column=0, sticky="w")
            com_type_combobox = ttk.Combobox(settings_frame, values=["RTU", "TCP"], state='readonly')
            com_type_combobox.grid(row=1, column=1, padx=5, pady=5)

            # Dynamic inputs based on Com Type, RTU settings shown by default
            settings_frame = ttk.Frame(settings_frame)
            settings_frame.grid(row=2, column=0, columnspan=2, sticky="ew")

            self.gen_settings.append({
                "panel_type": panel_type_combobox,
                "com_type": com_type_combobox,
                "settings_frame": settings_frame
            })

            # Update settings when Communication Type changes
            com_type_combobox.bind("<<ComboboxSelected>>", lambda event, i=i: self.on_com_type_selected(event, index=i))

    def on_com_type_selected(self, event, index):
        com_type = self.gen_settings[index-1]["com_type"].get()
        settings_frame = self.gen_settings[index-1]["settings_frame"]  # Direct access, assuming it's correctly stored
        if com_type == "RTU":
            self.gen_settings[index-1]["settings"] = self.show_rtu_settings(settings_frame)
        elif com_type == "TCP":
            self.gen_settings[index-1]["settings"] = self.show_tcp_settings(settings_frame)



    def show_rtu_settings(self, settings_frame):
        # Clear existing widgets in this frame
        for widget in settings_frame.winfo_children():
            widget.destroy()
        
        rtu_settings = {}  # Initialize rtu_settings dictionary

        # Define a function to update rtu_settings based on combobox changes
        def on_combobox_change(label_text, value):
            rtu_settings[label_text] = value

        # Slave Address Combobox
        ttk.Label(settings_frame, text="Slave Address:").grid(row=1, column=0, sticky='w')
        self.slave_address_combobox = ttk.Combobox(settings_frame, values=[str(n) for n in range(1, 31)], state='readonly')
        self.slave_address_combobox.set("10")  # Default value
        self.slave_address_combobox.grid(row=1, column=1, padx=5, pady=5)
        self.slave_address_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Slave Address", self.slave_address_combobox.get()))
    
        # Com Port Combobox
        ttk.Label(settings_frame, text="Com Port:").grid(row=2, column=0, sticky='w')
        self.com_port_combobox = ttk.Combobox(settings_frame, values=["COM1", "COM2", "COM3", "COM4"], state='readonly')
        self.com_port_combobox.set("COM1")  # Default value
        self.com_port_combobox.grid(row=2, column=1, padx=5, pady=5)
        self.com_port_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Com Port", self.com_port_combobox.get()))

        # Baudrate Combobox
        ttk.Label(settings_frame, text="Baudrate").grid(row=3, column=0, sticky='w')
        self.baudrate_combobox = ttk.Combobox(settings_frame, values=["9600", "19200", "115200"], state='readonly')
        self.baudrate_combobox.set("9600")  # Default value
        self.baudrate_combobox.grid(row=3, column=1, padx=5, pady=5)
        self.baudrate_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Baudrate", self.baudrate_combobox.get()))
        
        # Stopbit Combobox
        ttk.Label(settings_frame, text="Stopbit:").grid(row=4, column=0, sticky='w')
        self.stopbit_combobox = ttk.Combobox(settings_frame, values=["1", "1.5", "2"], state='readonly')
        self.stopbit_combobox.set("1")  # Default value
        self.stopbit_combobox.grid(row=4, column=1, padx=5, pady=5)
        self.stopbit_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Stopbit", self.stopbit_combobox.get()))
        
        # Parity Combobox
        ttk.Label(settings_frame, text="Parity:").grid(row=5, column=0, sticky='w')
        self.parity_combobox = ttk.Combobox(settings_frame, values=["None", "Odd", "Even"], state='readonly')
        self.parity_combobox.set("None")  # Default value
        self.parity_combobox.grid(row=5, column=1, padx=5, pady=5)
        self.parity_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Parity", self.parity_combobox.get()))
        
        # Databit Combobox
        ttk.Label(settings_frame, text="Databit:").grid(row=6, column=0, sticky='w')
        self.databit_combobox = ttk.Combobox(settings_frame, values=["7", "8"], state='readonly')
        self.databit_combobox.set("8")  # Default value
        self.databit_combobox.grid(row=6, column=1, padx=5, pady=5)
        self.databit_combobox.bind("<<ComboboxSelected>>", lambda event: on_combobox_change("Databit", self.databit_combobox.get()))

        rtu_settings = {
            "Slave Address": self.slave_address_combobox.get(),
            "Com Port": self.com_port_combobox.get(),
            "Baudrate": self.baudrate_combobox.get(),
            "Stopbit": self.stopbit_combobox.get(),
            "Parity": self.parity_combobox.get(),
            "Databit": self.databit_combobox.get()
        }
        return rtu_settings
    
    def show_tcp_settings(self, settings_frame):
        # Clear existing widgets in this frame
        for widget in settings_frame.winfo_children():
            widget.destroy()

        tcp_settings = {}  # Initialize tcp_settings dictionary

        def on_entry_change(label_text, value):
            tcp_settings[label_text] = value


        ttk.Label(settings_frame, text="IP Adress:").grid(row=1, column=0, sticky='w')
        self.ip_address_entry = ttk.Entry(settings_frame)
        self.ip_address_entry.insert(0,"192.168.1.110")
        self.ip_address_entry.grid(row=1, column=1, padx=5, pady=5)
        self.ip_address_entry.bind("<KeyRelease>", lambda event: on_entry_change("IP Address", self.ip_address_entry.get()))
            
        ttk.Label(settings_frame, text="Port:").grid(row=2, column=0, sticky='w')
        self.port_entry = ttk.Entry(settings_frame)
        self.port_entry.insert(0,"502")
        self.port_entry.grid(row=2, column=1, padx=5, pady=5)
        self.port_entry.bind("<KeyRelease>", lambda event: on_entry_change("Port", self.port_entry.get()))
            
        ttk.Label(settings_frame, text="Timeout:").grid(row=3, column=0, sticky='w')
        self.timeout_entry = ttk.Entry(settings_frame)
        self.timeout_entry.insert(0,"1000")
        self.timeout_entry.grid(row=3, column=1, padx=5, pady=5)
        self.timeout_entry.bind("<KeyRelease>", lambda event: on_entry_change("Timeout", self.timeout_entry.get()))
        
        tcp_settings = {
            "IP Address": self.ip_address_entry.get(),
            "Port": self.port_entry.get(),
            "Timeout": self.timeout_entry.get(),
        }
        return tcp_settings
    
    def create_file(self):
        # Example: Collecting input data
        barge_number = self.barge_number_entry.get()
        fjord_control_password = self.fjord_control_password_entry.get()
        send_interval = self.send_interval_entry.get()
        ems_plc_type = self.ems_plc_type_var.get()
        ess_plc_type = self.ess_plc_type_var.get()
        num_generators = int(self.num_generators_combobox.get())
        rtumodifications = {}
        tcpmodifications = {}
        next_tcp = 1  # Renamed from 'next' for clarity
            
        #Directory setup
        base_dir = f"Configs/FH{barge_number}"
        modules_dir = os.path.join(base_dir, "Modules")
        os.makedirs(modules_dir, exist_ok=True)
        
        #Copying the correct modules
        shutil.copy(f"Data/Modules/FCTags 3.0 {ems_plc_type}.nxbin", modules_dir)
        shutil.copy(f"Data/Modules/Effektberegninger.nxbin", modules_dir)
        shutil.copytree("Data/Modules/Data", os.path.join(modules_dir, "Data"))
       
        # Creating the Excel workbook
        config_excel_path = os.path.join(base_dir, "config.xlsx")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove the default sheet

        # Copy "Fjord Control 2" sheet and modify
        self.copy_sheet_from_template(
            os.path.join("Data", "FC2", "Fjord Control 2.xlsx"),wb,"Fjord Control 2",modifications={"C8": f"FH{barge_number}"}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "EMS", f"EMS modbus {ems_plc_type}.xlsx"),wb,"EMS modbus",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "EXT", "Ark1.xlsx"),wb,"Ark1",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "ESS", f"ESS modbus {ess_plc_type}.xlsx"),wb,"ESS modbus",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "Alarmlist", f"Alarmlist {num_generators}.xlsx"),wb,"AlarmList",modifications={}, namechange = None
        )
        self.copy_sheet_from_template(
            os.path.join("Data", "Fjord Control Online", f"Fjord Control Online {num_generators}.xlsx"),wb,"Fjord Control Online",modifications={"C5":f"bearer FH{barge_number}:{fjord_control_password}","C6":f"FH{barge_number}","C7":float(send_interval)}, namechange = None
        )
        
        num_generators = int(self.num_generators_combobox.get())  # Ensure this is defined above where it's used

        # Corrected loop
        for i in range(1,num_generators+1):
            # Use the index i-1 because list indexing starts at 0
            settings = self.gen_settings[i].get("settings")  # Corrected to use "settings" as a string key
            panel_type = self.gen_settings[i].get("panel_type").get()  # Assuming these are comboboxes, so use .get() to get the selected value
            com_type = self.gen_settings[i].get("com_type").get()  # Same as above

            # Now, i represents the generator number in a human-readable format
            print(f"Generator {i}: Panel Type: {panel_type}, Com Type: {com_type}, Settings: {settings}")



        """""
            com_port_column_map = {'COM1':'C','COM2':'F','COM3':'I','COM4':'L',}
            generator_ip_column_map = {'1':'I','2':'L','3':'O',}
            generator_ip_letter_column_map = {'1':'G','2':'J','3':'M',}

            if settings.get("com_type")== "RTU":
                com_port = settings.get("Com Port")
                if com_port in com_port_column_map:
                    base_column = com_port_column_map[com_port]
                    # Directly update the main rtumodifications dictionary
                    rtumodifications[f"{base_column}{5}"] = com_port
                    rtumodifications[f"{base_column}{6}"] = int(settings.get("Baudrate"))  # Default value if missing
                    rtumodifications[f"{base_column}{7}"] = float(settings.get("Stopbit"))
                    rtumodifications[f"{base_column}{8}"] = settings.get("Parity")
                    rtumodifications[f"{base_column}{9}"] = int(settings.get("Databit"))
            elif settings.get("com_type") == "TCP":
                Generator = f"Generator{i}"
                base_column = generator_ip_column_map[str(next_tcp)]
                base_column_letter = generator_ip_letter_column_map[str(next_tcp)]
                # Directly update the main tcpmodifications dictionary
                tcpmodifications[f"{base_column_letter}{5}"] = "Connector name"
                tcpmodifications[f"{base_column_letter}{6}"] = "Port no."
                tcpmodifications[f"{base_column_letter}{7}"] = "IP"
                tcpmodifications[f"{base_column}{5}"] = Generator
                tcpmodifications[f"{base_column}{6}"] = int(settings.get("Port"))  # Default value if missing
                tcpmodifications[f"{base_column}{7}"] = settings.get("IP Address")
                next_tcp += 1
        """""
        print(f"RTU Modifications: {rtumodifications}")
        print(f"TCP Modifications: {tcpmodifications}")
            
        # After collecting all modifications, copy 
        # the sheet template once
        if rtumodifications:
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "COM Ports.xlsx"), wb, "COM Ports", modifications=rtumodifications, namechange=None
            )
        else:
               # If no RTU settings found, copy the sheet without modifications
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "COM Ports.xlsx"), wb, "COM Ports", modifications={}, namechange=None
            )

        if tcpmodifications:
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "TCP Ports.xlsx"), wb, "TCP Ports", modifications=tcpmodifications, namechange=None
            )
        else:
               # If no TCP settings found, copy the sheet without modifications
            self.copy_sheet_from_template(
                os.path.join("Data", "COM", "TCP Ports.xlsx"), wb, "TCP Ports", modifications={}, namechange=None
            )
    
        self.copy_sheet_from_template(
           os.path.join("Data", "EXT", "Error values.xlsx"),wb,"Error values",modifications={}, namechange = None
        )

        # Save the new config
        self.save(config_excel_path, wb)
        
    def copy_sheet_from_template(self, template_path, new_wb, original_sheet_name, modifications=None, namechange=None, copy_style_from=("","")):
        # Load the template workbook
        template_wb = openpyxl.load_workbook(template_path, data_only=True)

        # Check if the original sheet exists in the template
        if original_sheet_name not in template_wb.sheetnames:
            print(f"Sheet '{original_sheet_name}' not found in {template_path}.")
            return

        # Get the template sheet
        template_sheet = template_wb[original_sheet_name]

        # Determine the new sheet name: use 'namechange' if provided, otherwise use the original sheet name
        new_sheet_name = namechange if namechange else original_sheet_name

        # Check if the new sheet name already exists in the target workbook to prevent overwriting
        if new_sheet_name in new_wb.sheetnames:
            print(f"Sheet name '{new_sheet_name}' already exists in the workbook. Please use a unique name.")
            return

        # Create a new sheet in the target workbook with the new sheet name
        new_sheet = new_wb.create_sheet(title=new_sheet_name)

        # Copy cells from template sheet to the new sheet, including values and styles
        for row in template_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)

        # Apply any modifications specified
        if modifications:
            for cell, value in modifications.items():
                new_sheet[cell].value = value

        # Copy merged cell ranges from the template sheet to the new sheet
        for merged_cell_range in template_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_cell_range))
        
        self.auto_adjust_column_widths_from_template(new_sheet)
        
        template_wb.close()
        
    def auto_adjust_column_widths_from_template(self, new_sheet):
        column_widths = {}
        default_width = 8.43

        for row in new_sheet.iter_rows():
            for cell in row:
                if not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Ignore merged cells for now
                        content_width = len(str(cell.value)) + 2 * 1.2
                        column_widths[cell.column_letter] = max(column_widths.get(cell.column_letter, default_width), content_width)

        for col_letter, width in column_widths.items():
            new_sheet.column_dimensions[col_letter].width = width
    
    def save(self, config_excel_path, wb):
        try:
            wb.save(config_excel_path)
            messagebox.showinfo("Success", "Config created successfully!", parent=self.root)
            self.root.destroy()  # Close the application after showing the success message
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save the config file.\n{e}", parent=self.root) 
                     
if __name__ == "__main__":
    # Ensure the current working directory is set to the script's or executable's directory
    root = tk.Tk()
    app = ExcelCreationToolGUI(root)
    root.mainloop()